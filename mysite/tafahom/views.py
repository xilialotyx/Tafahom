from django.utils import timezone
from django.views.generic.list import ListView
from .models import Tafahom, Vaam,ResPerTafahom,Organization,Instruction
from django.shortcuts import render,get_object_or_404
from jalali_date import datetime2jalali, date2jalali
from django.http import FileResponse
import openpyxl
from django.conf import settings
from openpyxl.styles import PatternFill
from pathlib import Path
from django.db.models import Q



class TafahomListView(ListView):
    model=Tafahom
    template_name = 'tafahom/list.html'  # Specify your own template name/location


    def get_queryset(self):
        tafahom_list = Tafahom.objects.all()
        search_param = self.request.GET.get('search_tafahom')

        if search_param:
            organization_obj = Organization.objects.filter(title__icontains=search_param)
            query = Q(num__icontains=search_param) | Q(organization__in=organization_obj)
            tafahom_list = Tafahom.objects.filter(query)


        return  tafahom_list

    # def get_context_data(self, **kwargs):
    #   # Call the base implementation first to get the context
    #   context = super(TafahomListView, self).get_context_data(**kwargs)
    #   # Create any data and add it to the context
    #   context['some_data'] = 'This is just some data'
    #   return context

def tafahom_details(request, tafahom_id):
    tafahom = get_object_or_404(Tafahom, pk=tafahom_id)
    instructions = Instruction.objects.filter(tafahom=tafahom_id)
    vaams = Vaam.objects.filter(instruction__in=instructions)
    res_type = ResPerTafahom.objects.filter(tafahom=tafahom_id)

    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        # بارگیری فایل اکسل
        wb = openpyxl.load_workbook(excel_file)
        # انتخاب ورک‌شیت
        ws = wb.active

        # تغییر رنگ سلول A1 به زرد
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        #ws['A1'].fill = yellow_fill

        for row in ws["A1":"E1"]:
            for cell in row:
                cell.fill = yellow_fill

        tafahom_id = request.POST["tafahom_id"]
        inst = Instruction.objects.filter(tafahom=tafahom_id)
        data_ws = []
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
            code_meli = row[0].value
            vaam = Vaam.objects.filter(code_meli=code_meli, instruction__in=inst).first()

            row[4].value=""
            if vaam is None:
                if tafahom.mablagh_limit and row[2].value > tafahom.mablagh_limit :
                    row[4].value += "-مبلغ وام بیشتر از حد تفاهم نامه"         
                
                if tafahom.modat_limit and row[3].value > tafahom.modat_limit :
                    row[4].value += "-مدت وام بیشتر از حد تفاهم نامه"
                
            else:
                # این کد ملی داخل این تفاهم نامه وام دریافت نکرده است
                row[4].value += "-کد ملی تکراری"
                if tafahom.mablagh_limit and row[2].value > tafahom.mablagh_limit :
                    row[4].value += "-مبلغ وام بیشتر از حد تفاهم نامه"         
                
                if tafahom.modat_limit and row[3].value > tafahom.modat_limit :
                    row[4].value += "-مدت وام بیشتر از حد تفاهم نامه"

            data_ws.append([code_meli,row[2].value,row[3].value,row[4].value])

        # ذخیره فایل اکسل تغییر یافته
        file_name:str='{}-{}.xlsx'.format(tafahom.num,datetime2jalali(timezone.now()).strftime('%Y%m%d-%H%M%S'))
        context = {
            'tafahom': tafahom,
            'res_type':res_type,
            'vaams': vaams,
            'ws':data_ws,
            'download_file_name':file_name 
        }
        wb.save(Path.joinpath(settings.BASE_DIR ,'tafahom', 'download',file_name) )
        

    else:
        context = {
            'tafahom': tafahom,
            'vaams': vaams,
            'res_type':res_type,
        }
    
    

    return render(request, 'tafahom/detail.html', context)



def download_file(request, file_name):
    file_path = Path.joinpath(settings.BASE_DIR ,'tafahom', 'download',file_name) 
    # اعتبارسنجی وجود فایل و انجام سایر عملیات مربوطه

    return FileResponse(open(file_path, 'rb'), as_attachment=True)